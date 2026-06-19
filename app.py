import streamlit as st
import pdfplumber
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

# Configuración de la página
st.set_page_config(page_title="PDF a Excel (Imágenes)", page_icon="📊", layout="centered")

st.title("📊 Convertidor de PDF a Excel (como Imágenes)")
st.write("Sube uno o varios archivos PDF. Cada PDF se convertirá en una pestaña de Excel, y dentro verás las páginas del PDF guardadas como imágenes fijas.")

# Componente para subir múltiples archivos
uploaded_files = st.file_uploader("Elige tus archivos PDF", type="pdf", accept_multiple_files=True)

if uploaded_files:
    st.success(f"¡{len(uploaded_files)} archivo(s) cargado(s) con éxito!")
    
    if st.button("Procesar y Generar Excel con Imágenes"):
        output = io.BytesIO()
        
        try:
            with st.spinner("Convirtiendo páginas de PDF a imágenes e insertándolas en Excel... Por favor espera."):
                # Creamos un libro de openpyxl nuevo
                wb = Workbook()
                # Eliminamos la hoja por defecto que trae openpyxl
                default_sheet = wb.active
                wb.remove(default_sheet)
                
                for uploaded_file in uploaded_files:
                    # Nombre de la hoja (máximo 31 caracteres permitido por Excel)
                    sheet_name = uploaded_file.name.replace(".pdf", "")[:31]
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # Abrir el PDF con pdfplumber
                    with pdfplumber.open(uploaded_file) as pdf:
                        row_offset = 2  # Fila inicial en Excel para colocar la primera página
                        
                        for i, page in enumerate(pdf.pages):
                            # Convertir la página del PDF a una imagen de alta resolución (150 DPI)
                            plumber_img = page.to_image(resolution=150)
                            pil_img = plumber_img.original  # Extraer formato nativo PIL (imagen)
                            
                            # Guardar la imagen temporalmente en memoria como PNG
                            img_buffer = io.BytesIO()
                            pil_img.save(img_buffer, format="PNG")
                            img_buffer.seek(0)
                            
                            # Convertir al objeto de imagen compatible con Excel
                            xl_img = OpenpyxlImage(img_buffer)
                            
                            # Colocar texto informativo en la columna A y la imagen en la columna B
                            ws[f"A{row_offset}"] = f"Página {i+1}"
                            cell_location = f"B{row_offset}"
                            
                            # Añadir la imagen directamente a la cuadrícula de Excel
                            ws.add_image(xl_img, cell_location)
                            
                            # Desplazar el marcador hacia abajo para que la siguiente página no se encima con la anterior
                            # Una página estándar a 150 DPI ocupa unas 65 filas en Excel de alto
                            row_offset += 65
                
                # Guardar el libro completo de Excel en nuestro buffer de memoria
                wb.save(output)
                output.seek(0)
                
                st.balloons()
                st.success("¡Archivo Excel con imágenes generado correctamente!")
                
                st.download_button(
                    label="📥 Descargar Archivo Excel",
                    data=output,
                    file_name="PDFs_Como_Imagenes.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        except Exception as e:
            st.error(f"Ocurrió un error al procesar los archivos: {e}")
